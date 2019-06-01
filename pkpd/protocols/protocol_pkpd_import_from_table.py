# **************************************************************************
# *
# * Authors:     Carlos Oscar Sorzano (info@kinestat.com)
# *
# * Kinestat Pharma
# *
# * This program is free software; you can redistribute it and/or modify
# * it under the terms of the GNU General Public License as published by
# * the Free Software Foundation; either version 2 of the License, or
# * (at your option) any later version.
# *
# * This program is distributed in the hope that it will be useful,
# * but WITHOUT ANY WARRANTY; without even the implied warranty of
# * MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# * GNU General Public License for more details.
# *
# * You should have received a copy of the GNU General Public License
# * along with this program; if not, write to the Free Software
# * Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA
# * 02111-1307  USA
# *
# *  All comments concerning this program package may be sent to the
# *  e-mail address 'info@kinestat.com'
# *
# **************************************************************************

import os
import sys

import pyworkflow.protocol.params as params
from .protocol_pkpd import ProtPKPD, addDoseToForm
from pkpd.objects import (PKPDExperiment, PKPDVariable, PKPDDose, PKPDVia,
                          PKPDSample)
from pyworkflow.utils import copyFile


class ProtPKPDImportFromTable(ProtPKPD):
    #--------------------------- DEFINE param functions --------------------------------------------

    def _defineParams(self, form):
        form.addSection('Input')
        inputFileHelp = "Specify a path to desired %s file.\n"%type
        inputFileHelp += "You may specify missing values with NA. You may also use LLOQ and ULOQ (Lower and Upper limit of quantification)\n"
        inputFileHelp += "to specify measurements that are below or above these limits"
        inputFileHelp += "If you use a CSV: the field separator must be a semicolon (;), decimal point must be a dot (.).\n"
        inputFileHelp += "If you use an Excel: you may have several sheets\n"
        inputFileHelp += "The first row must contain the name of the time variable and the name of the sample names\n"
        form.addParam('inputFile', params.PathParam,
                      label="File path", allowsNull=False, help=inputFileHelp)
        form.addParam('sheetName', params.StringParam, label="Sheet", default="", help="Only valid for Excel files. Leave empty for all sheets")
        form.addParam('title', params.StringParam, label="Title", default="My experiment")
        form.addParam('comment', params.StringParam, label="Comment", default="")
        form.addParam('tVar', params.TextParam, height=8, width=80, label="Time variable", default="t; h; numeric; time; Time variable")
        form.addParam('xVar', params.TextParam, height=8, width=80, label="Measurement variable", default="C; none; numeric; measurement; Measurement variable")


    #--------------------------- INSERT steps functions --------------------------------------------

    def _insertAllSteps(self):
        self._insertFunctionStep('createOutputStep',self.inputFile.get())

    #--------------------------- STEPS functions --------------------------------------------
    def addVar(self,experiment,line):
        tokens = line.split(';')
        if len(tokens) != 5:
            print("Skipping variable: ", line)
            return False
        varname = tokens[0].strip()
        experiment.variables[varname] = PKPDVariable()
        experiment.variables[varname].parseTokens(tokens)
        return True

    def readTables(self,fnTable):
        retval=[]
        if fnTable.endswith(".csv") or fnTable.endswith(".txt"):
            tableName=""
            # csv
            fh=open(fnTable,"r")
            lineNo=0
            allT=[]
            allSamples=[]
            for line in fh.readlines():
                tokens=line.split(";")
                if lineNo==0:
                    sampleNames=[]
                    tokenNo=0
                    for token in tokens:
                        if tokenNo>0:
                            strippedToken=token.strip()
                            if strippedToken!="":
                                sampleNames.append(strippedToken)
                        tokenNo+=1
                else:
                    allMeasurements=[]
                    tokenNo=0
                    for token in tokens:
                        if tokenNo==0:
                            allT.append(token.strip())
                        else:
                            strippedToken=token.strip()
                            if strippedToken!="":
                                allMeasurements.append(strippedToken)
                        tokenNo+=1
                    allSamples.append(allMeasurements)
                lineNo+=1
            fh.close()
            retval.append((tableName,sampleNames,allT,allSamples))
        else:
            # excel
            import openpyxl
            wb = openpyxl.load_workbook(fnTable)
            sheetNames=wb.get_sheet_names()
            for sheetName in sheetNames:
                if (self.sheetName.get()!="" and sheetName==self.sheetName.get()) or self.sheetName.get()=="":
                    tableName="" if len(sheetNames)==0 else "_"+sheetName
                    sampleNames = []
                    allT=[]
                    allSamples=[]
                    sheet = wb.get_sheet_by_name(sheetName)
                    for i in range(sheet.max_row):
                        if i>0:
                            allMeasurements=[]
                        for j in range(sheet.max_column):
                            cellValue = str(sheet.cell(row=i+1, column=j+1).value).strip()
                            if cellValue=="":
                                continue
                            if i==0 and j>=1:
                                sampleNames.append(cellValue)
                            elif i>0:
                                if j==0:
                                    allT.append(cellValue)
                                else:
                                    allMeasurements.append(cellValue)
                        if i>0:
                            allSamples.append(allMeasurements)
                    retval.append((tableName,sampleNames,allT,allSamples))
        return retval

    def createOutputStep(self, objId):
        fnFile = os.path.basename(self.inputFile.get())
        copyFile(self.inputFile.get(),self._getPath(fnFile))

        tables=self.readTables(self._getPath(fnFile))

        for tableName, sampleNames, allT, allMeasurements in tables:
            self.experiment = PKPDExperiment()
            self.experiment.general["title"] = self.title.get()
            self.experiment.general["comment"] = self.comment.get()

            # Add variables
            if not self.addVar(self.experiment,self.tVar.get()):
                raise Exception("Cannot process time variable")
            if not self.addVar(self.experiment,self.xVar.get()):
                raise Exception("Cannot process measurement variable")
            tvarName=self.tVar.get().split(';')[0].replace(';','')
            xvarName=self.xVar.get().split(';')[0].replace(';','')

            # Create the samples
            for sampleName in sampleNames:
                self.experiment.samples[sampleName] = PKPDSample()
                self.experiment.samples[sampleName].parseTokens([sampleName],self.experiment.variables,
                                                                self.experiment.doses, self.experiment.groups)
                self.experiment.samples[sampleName].addMeasurementPattern([sampleName,tvarName,xvarName])
                samplePtr = self.experiment.samples[sampleName]
                exec ("samplePtr.measurement_%s=%s" % (tvarName, allT))

            # Fill the samples
            for i in range(len(allT)):
                for j in range(len(sampleNames)):
                    samplePtr=self.experiment.samples[sampleNames[j]]
                    exec ('samplePtr.measurement_%s.append("%s")' % (xvarName, allMeasurements[i][j]))

            self.experiment.write(self._getPath("experiment%s.pkpd"%tableName))
            self.experiment._printToStream(sys.stdout)
            self._defineOutputs(**{"outputExperiment%s"%tableName: self.experiment})

    #--------------------------- INFO functions --------------------------------------------
    def _summary(self):
        return ["Input file: %s"%self.inputFile.get()]
