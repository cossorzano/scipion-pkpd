# **************************************************************************
# *
# * Authors:     Carlos Oscar Sorzano (info@kinestat.com)
# *
# * Kinestat Pharma
# *
# * This program is free software; you can redistribute it and/or modify
# * it under the terms of the GNU General Public License as published by
# * the Free Software Foundation; either version 2 of the License, or
# * (at your option) any later version.
# *
# * This program is distributed in the hope that it will be useful,
# * but WITHOUT ANY WARRANTY; without even the implied warranty of
# * MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# * GNU General Public License for more details.
# *
# * You should have received a copy of the GNU General Public License
# * along with this program; if not, write to the Free Software
# * Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA
# * 02111-1307  USA
# *
# *  All comments concerning this program package may be sent to the
# *  e-mail address 'info@kinestat.com'
# *
# **************************************************************************

import os
import sys

import pyworkflow.protocol.params as params
from .protocol_pkpd import ProtPKPD, addDoseToForm
from pkpd.objects import (PKPDExperiment, PKPDVariable, PKPDDose, PKPDVia,
                          PKPDSample)
from pyworkflow.utils import copyFile

WIDEFORMAT=0
LONGFORMAT=1

class ProtPKPDImportFromText(ProtPKPD):
    #--------------------------- DEFINE param functions --------------------------------------------

    def _defineParams(self, form, type):
        form.addSection('Input')
        inputFileHelp = "Specify a path to desired %s file.\n"%type
        inputFileHelp += "You may specify missing values with NA. You may also use LLOQ and ULOQ (Lower and Upper limit of quantification)\n"
        inputFileHelp += "to specify measurements that are below or above these limits"
        if type=="CSV":
            inputFileHelp += "The field separator must be a semicolon (;), decimal point must be a dot (.).\n"
            inputFileHelp += "The first row must contain the variable names and one of them must be SampleName which will serve as identifier."
        elif type=="ExcelTable":
            inputFileHelp += "The first column must be time, the rest of the columns must be measurements of each one of the individuals\n"
            inputFileHelp += "A column by each individual."
        form.addParam('inputFile', params.PathParam,
                      label="File path", allowsNull=False, help=inputFileHelp)
        if type=="ExcelTable":
            form.addParam('skipLines',params.IntParam, default=0, label='Skip lines',
                          help='Skip this amount of lines to reach the header line, the line with the variable names.')
            form.addParam('format',params.EnumParam, choices=["Wide format", "Long format"], default=0,
                          label='Excel format',
                          help='Wide format: 1st column is time, all other columns are the measurements\n'\
                               'Long format: 1st column is the individualID, then all columns are as described by the variables')
            form.addParam('header',params.StringParam, label='Header format', default='ID, t, Cp',
                          condition='format==1',
                          help='The ID column is compulsory, but it does not need to be the first one. You can skip columns with the keyword SKIP.')
        form.addParam('title', params.StringParam, label="Title", default="My experiment")
        form.addParam('comment', params.StringParam, label="Comment", default="")
        form.addParam('variables', params.TextParam, height=8, width=80, label="Variables", default="",
                      help="Structure: [Variable Name] ; [Units] ; [Type] ; [Role] ; [Comment]\n"\
                           "The variable name should have no space or special character\n"\
                           "Valid units are: h, mg, ug, ug/mL, ...\n"\
                           "Type is either numeric or text\n"\
                           "Role is either time, label or measurement\n"\
                           "The comment may be empty\n"\
                           "\nIt is important that there are three semicolons (none of them may be missing even if the comment is not present).\n"\
                           "Examples:\n"\
                           "t ; h ; numeric ; time ; \n"\
                           "Cp ; ug/mL ; numeric ; measurement ; plasma concentration\n"\
                           "weight ; g ; numeric; label ; weight of the animal\n"\
                           "sex ; none ; text ; label ; sex of the animal\n")
        form.addParam('vias', params.TextParam, height=8, width=80, label="Vias",
                      help="[ViaName]; [ViaType]; [tlag]; [bioavailability]"\
                       "Valid ViaTypes are: iv (intravenous), ev0 (extra-vascular order 0), ev1 (extra-vascular order 1), \n"\
                       "     ev01 (extra-vascular first order 0 and then order 1), evFractional (extra-vascular fractional order)\n"\
                       "     ev0tlag1 (extra-vascular first order 0 for a fraction F0, tlag1 and then order 1 for 1-F0)\n"\
                       "Optional parameters are tlag (e.g. tlag=0)\n"\
                       "   and bioavailability (e.g. bioavailability=0.8)\n"\
                       "Examples:\n"\
                       "Intravenous; iv\n"\
                       "Oral; ev1; tlag; bioavailability=1\n")
        addDoseToForm(form)
        form.addParam('dosesToSamples', params.TextParam, height=5, width=70, label="Assign doses to samples", default="",
                      help="Structure: [Sample Name] ; [DoseName1,DoseName2,...] \n"\
                           "The sample name should have no space or special character\n"\
                           "\nIt is important that there is one semicolon.\n"\
                           "Examples:\n"\
                           "FemaleRat1 ; Bolus0,Bolus1,Infusion0\n"\
                           "If left empty, all individuals are assigned to the first dose")


    #--------------------------- INSERT steps functions --------------------------------------------

    def _insertAllSteps(self):
        self._insertFunctionStep('createOutputStep',self.inputFile.get())

    #--------------------------- STEPS functions --------------------------------------------
    def readTextFile(self):
        pass

    def addSample(self,samplename,tokens):
        if not samplename in self.experiment.samples.keys():
            self.experiment.samples[samplename] = PKPDSample()
        self.experiment.samples[samplename].parseTokens(tokens, self.experiment.variables, self.experiment.doses,
                                                        self.experiment.groups)

    def createOutputStep(self, objId):
        fnFile = os.path.basename(self.inputFile.get())
        copyFile(self.inputFile.get(),self._getPath(fnFile))

        self.experiment = PKPDExperiment()
        self.experiment.general["title"]=self.title.get()
        self.experiment.general["comment"]=self.comment.get()

        ok = True

        # Read the variables
        self.listOfVariables = []
        for line in self.variables.get().replace('\n',';;').split(';;'):
            tokens = line.split(';')
            if len(tokens)!=5:
                print("Skipping variable: ",line)
                ok = False
                continue
            varname = tokens[0].strip()
            self.listOfVariables.append(varname)
            self.experiment.variables[varname] = PKPDVariable()
            self.experiment.variables[varname].parseTokens(tokens)

        # Read vias
        if self.vias.get():
            for line in self.vias.get().replace('\n',';;').split(';;'):
                if line!="":
                    tokens = line.split(';')
                    if len(tokens)<2:
                        print("Skipping via: ",line)
                        ok = False
                        continue
                    vianame = tokens[0].strip()
                    self.experiment.vias[vianame] = PKPDVia(ptrExperiment=self.experiment)
                    self.experiment.vias[vianame].parseTokens(tokens)

        # Read the doses
        if self.doses.get():
            for line in self.doses.get().replace('\n',';;').split(';;'):
                if line!="":
                    tokens = line.split(';')
                    if len(tokens)<5:
                        print("Skipping dose: ",line)
                        ok = False
                        continue
                    dosename = tokens[0].strip()
                    self.experiment.doses[dosename] = PKPDDose()
                    self.experiment.doses[dosename].parseTokens(tokens,self.experiment.vias)

        # Read the sample doses
        if self.dosesToSamples.get():
            for line in self.dosesToSamples.get().replace('\n',';;').split(';;'):
                try:
                    tokens = line.split(';')
                    samplename = tokens[0].strip()
                    if len(tokens)>1:
                        tokens[1]="dose="+tokens[1]
                    self.addSample(samplename,tokens)
                except Exception as e:
                    ok = False
                    print("Problem with line: ",line,str(e))

        if ok:
            # Read the measurements
            self.readTextFile()

            if self.dosesToSamples.get()=="" and self.experiment.doses: # There are doses but they are not assigned
               dosename = self.experiment.doses.keys()[0]
               for samplename in self.experiment.samples:
                   self.experiment.samples[samplename].doseList.append(dosename)

            self.experiment.write(self._getPath("experiment.pkpd"))
            self.experiment._printToStream(sys.stdout)
            self._defineOutputs(outputExperiment=self.experiment)

    #--------------------------- INFO functions --------------------------------------------
    def _summary(self):
        return ["Input file: %s"%self.inputFile.get()]

    def _validate(self):
        retval=[]
        if self.inputFile.get()=="" or self.inputFile.get is None:
            retval.append("There is no input file")
        else:
            if not os.path.exists(self.inputFile.get()):
                retval.append("The file %s does not exist"%self.inputFile.get())
        return retval


class ProtPKPDImportFromCSV(ProtPKPDImportFromText):
    """ Import experiment from CSV.\n
        You may use WebPlotDigitizer (https://apps.automeris.io/wpd) to generate this CSV, or Excel.
        Protocol created by http://www.kinestatpharma.com\n"""
    _label = 'import from csv'

    #--------------------------- DEFINE param functions --------------------------------------------

    def _defineParams(self, form):
        ProtPKPDImportFromText._defineParams(self,form,"CSV")
        form.addParam('noHeader', params.BooleanParam, default=False, label="No header",
                      help="There is no header in the CSV (e.g., this is the case of data taken from "
                      "WebPlotDigitizer, https://apps.automeris.io/wpd). If this is the case, "
                      "it is assumed that the columns of the CSV are in the same order as the variables are defined.")
        form.addParam('delimiter', params.StringParam, label='CSV Delimiter', default=";", help="Delimiter between fields in the CSV")
        form.addParam('sampleNameForm', params.StringParam, default="Individual", condition="noHeader",
                      label="Sample name")

    def readTextFile(self):
        fh=open(self.inputFile.get())
        lineNo = 1
        for line in fh.readlines():
            tokens = line.split(self.delimiter.get())
            if len(tokens)==0:
                continue
            if lineNo==1:
                listOfVariables=[]
                listOfSkips=[]
                if not self.noHeader:
                    iSampleName=-1
                    varNo=0
                    for token in tokens:
                        varName=token.strip()
                        if varName=="SampleName":
                            iSampleName=varNo
                        listOfVariables.append(varName)
                        listOfSkips.append(not (varName in self.experiment.variables))
                        varNo+=1
                    if iSampleName==-1:
                        raise Exception("Cannot find the SampleName in: %s\n"%line)
                else:
                    sampleName = self.sampleNameForm.get()
                    for line in self.variables.get().replace('\n', ';;').split(';;'):
                        tokens = line.split(';')
                        if len(tokens) != 5:
                            print("Skipping variable: ", line)
                            ok = False
                            continue
                        varname = tokens[0].strip()
                        listOfVariables.append(varname)
                        listOfSkips.append(False)
                    print("listOfVariables",listOfVariables)

            else:
                if len(tokens)!=len(listOfSkips):
                    print("Skipping line: %s"%line)
                    print("   It does not have the same number of values as the header")
                varNo = 0
                if not self.noHeader:
                    sampleName = tokens[iSampleName].strip()
                if not sampleName in self.experiment.samples:
                    self.addSample(sampleName,[sampleName])
                samplePtr=self.experiment.samples[sampleName]
                for skip in listOfSkips:
                    if not skip:
                        varName = listOfVariables[varNo]
                        varRole = self.experiment.variables[listOfVariables[varNo]].role
                        if varRole == PKPDVariable.ROLE_LABEL:
                            samplePtr.descriptors[listOfVariables[varNo]] = tokens[varNo]
                        else:
                            ok=True
                            if tokens[varNo]=="NA":
                                ok = (varRole != PKPDVariable.ROLE_TIME)
                            if ok:
                                if not hasattr(samplePtr,"measurement_%s"%varName):
                                    setattr(samplePtr,"measurement_%s"%varName,[])
                                    samplePtr.measurementPattern.append(varName)
                                exec("samplePtr.measurement_%s.append('%s')"%(varName,tokens[varNo].strip()))
                            else:
                                raise Exception("Time measurements cannot be NA")
                    varNo+=1
            lineNo+=1


class ProtPKPDImportFromExcel(ProtPKPDImportFromText):
    """ Import experiment from Excel.\n
        Protocol created by http://www.kinestatpharma.com\n"""
    _label = 'import from excel'

    #--------------------------- DEFINE param functions --------------------------------------------

    def _defineParams(self, form):
        ProtPKPDImportFromText._defineParams(self,form,"ExcelTable")

    def readTextFile(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.inputFile.get())
        sampleNames = []
        allT = []
        allSamples = []
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0]) # First sheet only
        i0=self.skipLines.get()
        if self.format.get()==WIDEFORMAT:
            for i in range(i0,sheet.max_row):
                if i-i0 > 0:
                    allMeasurements = []
                for j in range(sheet.max_column):
                    cellValue = str(sheet.cell(row=i + 1, column=j + 1).value).strip()
                    if cellValue == "":
                        continue
                    if i-i0 == 0 and j >= 1:
                        sampleName = cellValue
                        if sampleName[0] in "0123456789":
                            sampleName="d"+sampleName
                        sampleNames.append(sampleName)
                    elif i-i0 > 0:
                        if j == 0:
                            allT.append(cellValue)
                        else:
                            allMeasurements.append(cellValue)
                if i-i0 > 0:
                    allSamples.append(allMeasurements)

            tvarName = None
            xvarName = None
            for varName in self.experiment.variables:
                if self.experiment.variables[varName].role == PKPDVariable.ROLE_TIME:
                    tvarName = varName
                elif self.experiment.variables[varName].role == PKPDVariable.ROLE_MEASUREMENT:
                    xvarName = varName

            for sampleName in sampleNames:
                self.addSample(sampleName,[sampleName])
                samplePtr=self.experiment.samples[sampleName]
                samplePtr.addMeasurementPattern([sampleName, tvarName, xvarName])
                exec ("samplePtr.measurement_%s=%s" % (tvarName, allT))

            for i in range(len(allT)):
                for j in range(len(sampleNames)):
                    samplePtr=self.experiment.samples[sampleNames[j]]
                    exec ('samplePtr.measurement_%s.append("%s")' % (xvarName, allSamples[i][j]))

        elif self.format.get()==LONGFORMAT:
            headerFormat=[token.strip() for token in self.header.get().split(',')]
            if len(headerFormat)!=sheet.max_column:
                raise Exception("You have specified %d columns in the header format, but there are %d columns"\
                                %(len(headerFormat),sheet.max_column))
            if not "ID" in headerFormat:
                raise Exception("Cannot find ID in the header format")
            idCol = headerFormat.index("ID")

            keepCols=[]
            col=0
            measurementPattern=[]
            for colName in headerFormat:
                if colName != "SKIP" and colName !="ID":
                    keepCols.append(col)
                    measurementPattern.append(colName)
                col+=1
            print(keepCols)

            allSamples={}
            for i in range(i0+1,sheet.max_row):
                line=[]
                for j in range(sheet.max_column):
                    cellValue = str(sheet.cell(row=i + 1, column=j + 1).value).strip()
                    line.append(cellValue)
                sampleName = line[idCol]
                if sampleName[0] in "0123456789":
                    sampleName="d"+sampleName
                if not sampleName in allSamples:
                    allSamples[sampleName]=[]
                    for j in range(len(keepCols)):
                        allSamples[sampleName].append([])
                jidx=0
                for j in keepCols:
                    allSamples[sampleName][jidx].append(line[j])
                    jidx+=1

            for sampleName in allSamples:
                self.addSample(sampleName,[sampleName])
                samplePtr=self.experiment.samples[sampleName]
                samplePtr.addMeasurementPattern([sampleName]+measurementPattern)
                for jidx in range(len(keepCols)):
                    exec ("samplePtr.measurement_%s=%s" % (measurementPattern[jidx], allSamples[sampleName][jidx]))


def getSampleNamesFromCSVfile(fnCSV, delimiter=';'):
    sampleNames = []
    fh=open(fnCSV)
    lineNo = 1
    for line in fh.readlines():
        tokens = line.split(delimiter)
        if len(tokens)==0:
            continue
        if lineNo==1:
            iSampleName=-1
            varNo = 0
            for token in tokens:
                varName=token.strip()
                if varName=="SampleName":
                    iSampleName=varNo
                    break
                varNo += 1
            if iSampleName==-1:
                fh.close()
                return
        else:
            if len(tokens)>iSampleName:
                sampleName = tokens[iSampleName]
                if not sampleName in sampleNames:
                    sampleNames.append(sampleName.strip())
        lineNo+=1
    fh.close()
    return sampleNames


def getVarNamesFromCSVfile(fnCSV, delimiter=';'):
    varNames = []
    fh=open(fnCSV)
    for line in fh.readlines():
        tokens = line.split(delimiter)
        if len(tokens)==0:
            continue
        for token in tokens:
            varNames.append(token.strip())
        break
    fh.close()
    return varNames
