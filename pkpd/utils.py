# **************************************************************************
# *
# * Authors:     Carlos Oscar Sorzano (info@kinestat.com)
# *
# * Kinestat Pharma
# *
# * This program is free software; you can redistribute it and/or modify
# * it under the terms of the GNU General Public License as published by
# * the Free Software Foundation; either version 2 of the License, or
# * (at your option) any later version.
# *
# * This program is distributed in the hope that it will be useful,
# * but WITHOUT ANY WARRANTY; without even the implied warranty of
# * MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# * GNU General Public License for more details.
# *
# * You should have received a copy of the GNU General Public License
# * along with this program; if not, write to the Free Software
# * Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA
# * 02111-1307  USA
# *
# *  All comments concerning this program package may be sent to the
# *  e-mail address 'info@kinestat.com'
# *
# **************************************************************************
"""
PKPD functions
"""
import copy
from itertools import izip
import numpy as np
import math
from scipy.interpolate import InterpolatedUnivariateSpline, pchip_interpolate
import time
import hashlib
from os.path import (exists, splitext, getmtime)
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

def parseRange(auxString):
    if auxString=="":
        return None
    elif auxString.startswith('['):
        auxString=auxString.replace('[','')
        auxString=auxString.replace(']','')
        tokens=auxString.split(',')
        auxArray = np.array(tokens, dtype='float')
        return auxArray.astype(np.float)
    elif ':' in auxString:
        tokens=auxString.split(':')
        if len(tokens)!=3:
            raise Exception("The X evaluation string is not well formatted: %s"%auxString)
        fromValue = float(tokens[0])
        step= float(tokens[1])
        toValue = float(tokens[2])
        return np.arange(fromValue,toValue,step)

def find_nearest(array,value):
    idx = np.searchsorted(array, value, side="left")
    if idx > 0 and (idx == len(array) or math.fabs(value - array[idx-1]) < math.fabs(value - array[idx])):
        return idx-1
    else:
        return idx

def getMD5String(fn):
    fnTime = time.strftime("%Y-%m-%d-%M:%S\n",time.gmtime(getmtime(fn)))
    return hashlib.md5(fnTime+open(fn, 'rb').read()).hexdigest()

def writeMD5(fn):
    if not exists(fn):
        return
    fnMD5=splitext(fn)[0]+".md5"
    fh=open(fnMD5,"w")
    fh.write("%s\n"%getMD5String(fn))
    fh.write("%s\n"%fn)
    fh.close()

def verifyMD5(fn):
    if fn is None or not exists(fn):
        return True
    if fn.endswith(".md5"):
        fnMD5=fn
    else:
        fnMD5=splitext(fn)[0]+".md5"
        if not exists(fnMD5):
            return False
    fh=open(fnMD5,"r")
    md5StringFile=fh.readline().strip()
    fnFile=fh.readline().strip()
    fh.close()
    if fn.endswith(".md5"):
        fn=fnFile
    return getMD5String(fn)==md5StringFile

def uniqueFloatValues(x,y, TOL=-1):
    xp=np.asarray(x,dtype=np.float64)
    yp=np.asarray(y,dtype=np.float64)
    idx = np.logical_not(np.logical_or(np.isnan(xp),np.isnan(yp)))
    xp=xp[idx]
    yp=yp[idx]
    if TOL<0:
        TOL = np.max(xp.flat) / (xp.size*1e3)
    A = np.zeros((xp.size,), dtype=[('x','f8'),('y','f8')])
    A['x']=xp.flat
    A['y']=yp.flat
    idx = np.argsort(A,order=('x','y')) # A is constructed for solving ties
    d = np.append(1, np.diff(xp.flat[idx]))
    #for i in range(0,d.size):
    #    print("i=",i,"xp[i]=",xp[i],"d[i]=",d[i],"yp[i]=",yp[i],"idx[i]",idx[i])
    xunique = xp.flat[idx[d > TOL]]
    yunique = yp.flat[idx[d > TOL]]
    return xunique,yunique

def uniqueFloatValues2(x1,x2,y):
    x1p=np.asarray(x1,dtype=np.float64)
    x2p=np.asarray(x2,dtype=np.float64)
    yp=np.asarray(y,dtype=np.float64)

    TOL1 = np.max(x1p.flat) / 1e5
    idx1 = np.argsort(x1p.flat)
    d1 = np.append(True, np.diff(x1p.flat[idx1]))

    TOL2 = np.max(x2p.flat) / 1e5
    idx2 = np.argsort(x1p.flat)
    d2 = np.append(True, np.diff(x2p.flat[idx2]))

    if np.sum(d1>TOL1)>np.sum(d2>TOL2):
        takeIdx=idx1[d1>TOL1]
    else:
        takeIdx=idx2[d2>TOL2]
    x1unique = x1p.flat[takeIdx]
    x2unique = x2p.flat[takeIdx]
    yunique = yp.flat[takeIdx]
    return x1unique,x2unique,yunique

def twoWayUniqueFloatValues(x,y):
    y1,x1=uniqueFloatValues(y,x)
    x2,y2=uniqueFloatValues(x1,y1)
    return x2,y2

def calculateAUC0t(t, C):
    # Make sure that the (0,0) sample is present
    AUC0t = np.zeros(t.shape[0])
    for idx in range(1,t.shape[0]):
        dt = (t[idx]-t[idx-1])
        if C[idx]>=C[idx-1]: # Trapezoidal in the raise
            AUC0t[idx] = AUC0t[idx-1]+0.5*dt*(C[idx-1]+C[idx])
        else: # Log-trapezoidal in the decay
            if C[idx-1]>0 and C[idx]>0:
                decrement = C[idx-1]/C[idx]
                K = math.log(decrement)
                B = K/dt
                AUC0t[idx] = AUC0t[idx-1]+dt*(C[idx-1]-C[idx])/K
            else:
                AUC0t[idx] = AUC0t[idx-1]
    return AUC0t

def upper_tri_masking(A):
    # Extract the upper triangular matrix without the diagonal
    r = np.arange(A.shape[0])
    c = np.arange(A.shape[1])
    mask = r[:,None] < r
    return A[mask]

def parseOperation(operation):
    varList = []
    idx0 = operation.find("$(")
    while idx0 >= 0:
        idxF = operation.find(")", idx0)
        varName = operation[(idx0 + 2):idxF]
        if not varName in varList:
            varList.append(varName)
        idx0 = operation.find("$(", idxF + 1)

    coeffList = []
    idx0 = operation.find("$[")
    while idx0 >= 0:
        idxF = operation.find("]", idx0)
        varName = operation[(idx0 + 2):idxF]
        if not varName in varList:
            coeffList.append(varName)
        idx0 = operation.find("$[", idxF + 1)

    parsedOperation = copy.copy(operation)
    for varName in varList:
        exec ("parsedOperation=parsedOperation.replace('$(%s)','%s')" % (varName, varName))
    for varName in coeffList:
        exec ("parsedOperation=parsedOperation.replace('$[%s]','%s')" % (varName, varName))
    return parsedOperation, varList, coeffList

def excelWriteRow(msgList, workbook, row, col=1, sheetName="", bold=False):
    currentCol=col
    if sheetName!="":
        if not sheetName in workbook.sheetnames:
            workbook.create_sheet(sheetName)
        sheet=workbook[sheetName]
    else:
        sheet=workbook.active
    if type(msgList)!=list:
        msgList2=[msgList]
    else:
        msgList2=msgList
    for msg in msgList2:
        c = sheet.cell(row=row, column=currentCol)
        try:
            if isinstance(msg,bool):
                c.value = str(msg)
            else:
                c.value = msg
            if bold:
                c.font = Font(bold=True)
        except:
            print("Cannot convert the cell row=%d, col=%d"%(row,col),msg)
        currentCol+=1

def excelFillCells(workbook, row, col0=1, colF=10, sheetName="", fillColor="54B948"):
    if sheetName!="":
        if not sheetName in workbook.sheetnames:
            workbook.create_sheet(sheetName)
        sheet=workbook[sheetName]
    else:
        sheet=workbook.active
    for col in range(col0,colF+1):
        c = sheet.cell(row=row, column=col)
        c.fill = PatternFill(bgColor=fillColor, fill_type="solid")


def as_text(value):
    if value is None:
        return ""
    return str(value)

def excelAdjustColumnWidths(workbook, sheetName=""):
    if sheetName!="":
        if not sheetName in workbook.sheetnames:
            workbook.create_sheet(sheetName)
        sheet=workbook[sheetName]
    else:
        sheet=workbook.active

    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

def computeXYmean(XYlist, Nxsteps=300):
    xmin = 1e38
    xmax = -1e38
    for x,_ in XYlist:
        xmin = min(xmin, np.min(x))
        xmax = max(xmax, np.max(x))

    dataDict = {}  # key will be x values
    xrange = np.arange(xmin, xmax, (xmax - xmin) / Nxsteps)
    for xValues, yValues in XYlist:
        xValuesUnique, yValuesUnique = twoWayUniqueFloatValues(xValues, yValues)
        xUniqueMin=np.min(xValuesUnique)
        xUniqueMax=np.max(xValuesUnique)
        B = InterpolatedUnivariateSpline(xValuesUnique, yValuesUnique, k=1)
        xrangei=xrange[np.logical_and(xrange>=xUniqueMin,xrange<=xUniqueMax)]
        yrange = B(xrangei)
        for x, y in izip(xrange, yrange):
            if x in dataDict:
                dataDict[x].append(y)
            else:
                dataDict[x] = [y]

    sortedX = sorted(dataDict.keys())

    # We will store five values (min, 25%, 50%, 75%, max)
    # for each of the time entries computed
    # percentileList = [0, 25, 50, 75, 100]
    # Y = np.zeros((len(sortedX), 5))
    # for i, t in enumerate(sortedX):
    #    Y[i, :] = np.percentile(dataDict[t], percentileList)

    Y = np.zeros(len(sortedX))
    for i, t in enumerate(sortedX):
        Y[i] = np.mean(dataDict[t])

    return sortedX, Y

def smoothPchip(x,y):
    d=np.insert(np.diff(y),0,0)
    d[d<0]=0
    ypos=np.cumsum(d)
    ypos=ypos*np.sum(y)/np.sum(ypos)
    return ypos
    #xunique, yunique = uniqueFloatValues(x,ypos)
    #yInterpolated = pchip_interpolate(xunique,yunique,x)
    #return yInterpolated